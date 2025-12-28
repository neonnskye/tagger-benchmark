import json
import logging
import re
from pathlib import Path

import yaml
from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Border, Side, PatternFill

import config

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

BENCHMARK_TEMPLATE_PATH = "templates/benchmark.xlsx"
BENCHMARK_OUTPUT_PATH = "results/results.xlsx"

BENCHMARK_START_COLUMN = 5
BENCHMARK_START_ROW = 1

MASTER_TAG_LIST_PATH = "tags/master.yaml"
USER_PROMPT_PATH = "outputs/prompt/user.md"

RESPONSE_OUTPUT_DIR = Path("outputs")


class Styles:
    bold_font = Font(bold=True)
    thin_side = Side(style="thin")
    gray_fill = PatternFill(
        fill_type="solid",
        start_color="BFBFBF",
        end_color="BFBFBF"
    )
    blue_fill = PatternFill(
        fill_type="solid",
        start_color="0070C0",
        end_color="0070C0"
    )
    yellow_fill = PatternFill(
        fill_type="solid",
        start_color="FFFF00",
        end_color="FFFF00"
    )
    red_fill = PatternFill(
        fill_type="solid",
        start_color="C00000",
        end_color="C00000"
    )
    all_borders = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )


def make_json_safe(obj):
    if isinstance(obj, set):
        return sorted(obj)  # or list(obj)
    if isinstance(obj, dict):
        return {k: make_json_safe(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [make_json_safe(v) for v in obj]
    return obj


class Benchmark:
    def __init__(self):
        self.master_tag_list = set()
        self.all_model_names = set()
        self.num_total_models = 0
        self.load_references()

        self.response_file_paths = {}
        self.find_response_file_paths()

        self.responses = {}
        self.parse_responses()

        self.image_list = list(config.IMAGES_CAPTIONS.keys())
        self.image_caption_list = []
        self.image_tag_list = []

        self.wb = load_workbook(BENCHMARK_TEMPLATE_PATH)
        self.ws = self.wb["Benchmark"]
        self.prep_result_sheet()

        self.run_analysis()
        self.wb.save(BENCHMARK_OUTPUT_PATH)
        logger.info(f"Benchmark results saved to '{BENCHMARK_OUTPUT_PATH}'")

    def load_references(self):
        def extract_master_tag_list(obj):
            if isinstance(obj, dict):
                for value in obj.values():
                    extract_master_tag_list(value)
            elif isinstance(obj, list):
                for item in obj:
                    if isinstance(item, str) and item.startswith('#'):
                        self.master_tag_list.add(item[1:])
                    else:
                        extract_master_tag_list(item)

        with open(MASTER_TAG_LIST_PATH, 'r', encoding='utf-8') as master_tag_file:
            master_tag_data = yaml.safe_load(master_tag_file)
        extract_master_tag_list(master_tag_data)

        self.all_model_names = {model.split("/", 1)[1] for model in config.MODEL_LIST}
        self.num_total_models = len(self.all_model_names)

    def find_response_file_paths(self):
        logging.info("Scanning responses...")
        found_file_paths = [p for p in RESPONSE_OUTPUT_DIR.iterdir() if p.is_file()]
        for file_path in found_file_paths:
            for model_name in config.MODEL_LIST:
                short_name = model_name.split("/", 1)[1]  # Extract name after provider/
                if file_path.stem == short_name:  # Use .stem for exact match without extension
                    self.response_file_paths[model_name] = file_path
        logging.info(f"Found {len(self.response_file_paths)} valid response files")

    def parse_responses(self):
        tags_parsed = 0
        logging.info("Parsing responses...")
        for response_file_path in self.response_file_paths.values():
            model_name = response_file_path.name[:-3]

            with open(response_file_path, "r", encoding="utf-8") as response_file:
                response_raw = response_file.read()

            match = re.search(r"({[\W\w]+})", response_raw)
            if not match:
                logging.warning(f"Skipped invalid JSON response: '{response_file_path}'...")
            response = json.loads(match.group(1))

            image_results = response["results"]
            for image_idx, image_result in enumerate(image_results):
                image_key = f"image_{image_idx + 1}"
                tags = {tag.lstrip("#") for tag in image_result["tags"]}
                tags_parsed += len(tags)

                if image_key not in self.responses:
                    self.responses[image_key] = {
                        "by_model": {},
                        "tag_index": {},
                    }

                self.responses[image_key]["by_model"][model_name] = tags

                tag_index = self.responses[image_key]["tag_index"]
                for tag in tags:
                    if tag not in tag_index:
                        tag_index[tag] = set()
                    tag_index[tag].add(model_name)

        logger.info(f"Parsed a total of {tags_parsed} unique tags")

        # For debugging
        # with open("out.json", "w", encoding="utf-8") as out_file:
        #     safe_json = make_json_safe(self.responses)
        #     json.dump(safe_json, out_file, indent=2)

    def prep_result_sheet(self):
        with open(USER_PROMPT_PATH, "r", encoding="utf-8") as user_prompt_file:
            user_prompt = user_prompt_file.read()
            self.image_caption_list = re.findall(r"Caption\n\n```\n(.*)", user_prompt)
            self.image_tag_list = re.findall(r"WD14 tags\n\n```\n(.*)", user_prompt)

        num_images = len(self.image_list)
        num_models = len(self.all_model_names)
        rows_per_block = 2 + num_models  # caption + tags + models

        for i, row in enumerate(range(1, 1 + rows_per_block * num_images, rows_per_block)):
            self.ws[f"A{row}"] = self.image_caption_list[i]
            self.ws[f"A{row + 1}"] = self.image_tag_list[i]

            # Insert image
            image_path = self.image_list[i]

            with Image.open(image_path) as img:
                img_width, img_height = img.size

            col_width = self.ws.column_dimensions["B"].width

            scaling_factor = img_width / (col_width * 7)
            logger.debug(f"scaling_factor: {img_width} / ({col_width} * 7) = {scaling_factor}")

            xl_img = XLImage(image_path)
            xl_img.width = int(img_width / scaling_factor)
            xl_img.height = int(img_height / scaling_factor)

            self.ws.add_image(xl_img, f"B{row}")

    def run_analysis(self):
        logger.info("Running benchmark...")

        # To place tag data dynamically
        model_name_indexes = {}
        for model_name in self.all_model_names:
            matches = [i for i, s in enumerate(config.MODEL_LIST) if re.search(rf"/({model_name})$", s)]
            model_name_indexes[model_name] = matches[0] + 1
        logger.debug(f"model_name_indexes: {model_name_indexes}")

        tag_row = BENCHMARK_START_ROW
        for image_name, image_data in self.responses.items():
            tag_column = BENCHMARK_START_COLUMN
            for tag_name, models_selected in image_data["tag_index"].items():
                # Tag labels
                label_cell = self.ws.cell(row=tag_row, column=tag_column)
                label_cell.value = tag_name
                label_cell.font = Styles.bold_font
                label_cell.border = Styles.all_borders

                num_models_selected = len(models_selected)
                # num_models_not_selected = self.num_total_models - num_models_selected
                # models_not_selected = self.all_model_names - models_selected

                gold_tag = False
                if (num_models_selected / self.num_total_models) >= 0.7:
                    gold_tag = True

                invalid_tag = False
                if tag_name not in self.master_tag_list:
                    invalid_tag = True

                for model_name, index in model_name_indexes.items():
                    data_row = tag_row + index
                    data_cell = self.ws.cell(row=data_row, column=tag_column)
                    if model_name in models_selected:
                        if gold_tag:
                            data_cell.fill = Styles.blue_fill
                            data_cell.value = config.Scoring.AGREEMENT_HIT_BONUS
                        else:
                            data_cell.fill = Styles.gray_fill
                        if invalid_tag:
                            data_cell.fill = Styles.red_fill
                            data_cell.value = config.Scoring.HALLUCINATION_PENALTY
                    else:
                        if gold_tag:
                            data_cell.fill = Styles.yellow_fill
                            data_cell.value = config.Scoring.AGREEMENT_MISS_PENALTY
                    data_cell.border = Styles.all_borders

                tag_column += 1
            tag_row += (len(self.all_model_names) + 2)


if __name__ == "__main__":
    Benchmark()
